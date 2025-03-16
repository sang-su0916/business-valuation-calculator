import streamlit as st

# í˜ì´ì§€ ì„¤ì • - ë°˜ë“œì‹œ ì²« ë²ˆì§¸ Streamlit ëª…ë ¹ì–´ë¡œ ìœ„ì¹˜í•´ì•¼ í•¨
st.set_page_config(
    page_title="ë¹„ìƒì¥ì£¼ì‹ í‰ê°€",
    page_icon="ğŸ“Š",
    layout="wide",
    initial_sidebar_state="expanded"
)

import pandas as pd
import numpy as np
import locale
from datetime import datetime
import io
import base64
import re
import json

# FPDF ë¼ì´ë¸ŒëŸ¬ë¦¬ ì¶”ê°€ (PDF ìƒì„±ìš©)
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    try:
        # pipìœ¼ë¡œ ì„¤ì¹˜ ì‹œë„
        import subprocess
        subprocess.check_call(['pip', 'install', 'fpdf'])
        from fpdf import FPDF
        FPDF_AVAILABLE = True
    except:
        FPDF_AVAILABLE = False

# íŒŒì¼ ì²˜ë¦¬ ë¼ì´ë¸ŒëŸ¬ë¦¬
try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# ìˆ«ì í˜•ì‹í™”ë¥¼ ìœ„í•œ ë¡œì¼€ì¼ ì„¤ì •
try:
    locale.setlocale(locale.LC_ALL, 'ko_KR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'ko_KR')
    except:
        locale.setlocale(locale.LC_ALL, '')

# í˜ì´ì§€ ìŠ¤íƒ€ì¼ë§
st.markdown("""
<style>
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1000px;
    }
    .stNumberInput input {
        font-weight: bold;
    }
    .stExpander {
        border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
    }
    .amount-display {
        color: #0066cc;
        font-weight: bold;
        padding: 5px 0;
    }
    .field-description {
        color: #666;
        font-style: italic;
        font-size: 0.85rem;
        margin-top: 0.2rem;
        margin-bottom: 0.5rem;
    }
    .section-header {
        font-weight: bold;
        font-size: 1.1rem;
        margin-bottom: 0.5rem;
        color: #333;
    }
    .upload-section {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 8px;
        margin-bottom: 15px;
    }
    .download-section {
        background-color: #f0f7fb;
        padding: 15px;
        border-radius: 8px;
        margin-top: 20px;
    }
    .number-input-container {
        display: flex;
        align-items: center;
        margin-bottom: 5px;
    }
    .number-input-container .stNumberInput {
        flex-grow: 1;
        margin-right: 10px;
    }
    .reset-button {
        color: #dc3545;
        border: 1px solid #dc3545;
        padding: 3px 8px;
        border-radius: 4px;
        cursor: pointer;
        font-weight: bold;
        font-size: 12px;
        background-color: white;
    }
    .reset-button:hover {
        background-color: #dc3545;
        color: white;
    }
    .unit-selector {
        margin-bottom: 5px;
    }
</style>
""", unsafe_allow_html=True)

# ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™”
if 'eval_date' not in st.session_state:
    st.session_state.eval_date = datetime.now().date()
if 'company_name' not in st.session_state:
    st.session_state.company_name = "ì£¼ì‹íšŒì‚¬ ì—ì´ë¹„ì”¨"
if 'total_equity' not in st.session_state:
    st.session_state.total_equity = 1000000000
if 'net_income1' not in st.session_state:
    st.session_state.net_income1 = 386650000
if 'net_income2' not in st.session_state:
    st.session_state.net_income2 = 163401000
if 'net_income3' not in st.session_state:
    st.session_state.net_income3 = 75794000
if 'shares' not in st.session_state:
    st.session_state.shares = 4000
if 'owned_shares' not in st.session_state:
    st.session_state.owned_shares = 2000
if 'share_price' not in st.session_state:
    st.session_state.share_price = 5000
if 'interest_rate' not in st.session_state:
    st.session_state.interest_rate = 10  # í™˜ì›ìœ¨ 10%ë¡œ ê³ ì •
if 'evaluation_method' not in st.session_state:
    st.session_state.evaluation_method = "ì¼ë°˜ë²•ì¸"
if 'stock_value' not in st.session_state:
    st.session_state.stock_value = None
if 'evaluated' not in st.session_state:
    st.session_state.evaluated = False
if 'shareholders' not in st.session_state:
    st.session_state.shareholders = [
        {"name": "ëŒ€í‘œì´ì‚¬", "shares": 2000},
        {"name": "", "shares": 0},
        {"name": "", "shares": 0},
        {"name": "", "shares": 0},
        {"name": "", "shares": 0}
    ]
if 'shareholder_count' not in st.session_state:
    st.session_state.shareholder_count = 1
    
# ë‹¨ìœ„ ì˜µì…˜ ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™”
if 'total_equity_unit' not in st.session_state:
    st.session_state.total_equity_unit = "ì›"
if 'net_income1_unit' not in st.session_state:
    st.session_state.net_income1_unit = "ì›"
if 'net_income2_unit' not in st.session_state:
    st.session_state.net_income2_unit = "ì›"
if 'net_income3_unit' not in st.session_state:
    st.session_state.net_income3_unit = "ì›"
if 'share_price_unit' not in st.session_state:
    st.session_state.share_price_unit = "ì›"

# ë‹¨ìœ„ ë³€í™˜ í•¨ìˆ˜
def convert_to_base_unit(value, unit):
    if unit == "ì²œì›":
        return value * 1000
    return value

# ë‹¨ìœ„ì— ë§ê²Œ í‘œì‹œ í˜•ì‹ ë³€í™˜
def format_by_unit(value, unit):
    if unit == "ì²œì›":
        return format_number(value // 1000)
    return format_number(value)

# ìˆ«ì í˜•ì‹í™” í•¨ìˆ˜
def format_number(num, in_thousands=False):
    try:
        if in_thousands:
            # ì²œì› ë‹¨ìœ„ë¡œ ë³€í™˜ (1,000,000,000ì› -> 1,000,000ì²œì›)
            return "{:,}".format(int(num) // 1000) + "ì²œ"
        else:
            return "{:,}".format(int(num))
    except:
        return str(num)

# ì…ë ¥ í•„ë“œ ë¦¬ì…‹ í•¨ìˆ˜
def reset_field(field_key):
    st.session_state[field_key] = 0

# ì •ê·œì‹ íŒ¨í„´ìœ¼ë¡œ ë°ì´í„° ì°¾ê¸°
def find_with_patterns(text, patterns):
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # ìˆ«ìë§Œ ì¶”ì¶œ
            for match in matches:
                num_str = re.sub(r'[^\d]', '', match)
                if num_str:
                    return int(num_str)
    return None

# íŒŒì¼ì—ì„œ ë°ì´í„° ì¶”ì¶œ
def extract_data_from_file(uploaded_file):
    extracted_data = {}
    
    # íŒŒì¼ í™•ì¥ì í™•ì¸
    file_ext = uploaded_file.name.split('.')[-1].lower()
    
    # í…ìŠ¤íŠ¸ ì¶”ì¶œ
    text_content = ""
    
    # Excel íŒŒì¼ ì²˜ë¦¬
    if file_ext in ['xlsx', 'xls'] and XLSX_AVAILABLE:
        try:
            df = pd.read_excel(uploaded_file)
            # DataFrameì„ í…ìŠ¤íŠ¸ë¡œ ë³€í™˜
            text_content = df.to_string()
            
            # ì—´ ì´ë¦„ìœ¼ë¡œ ë°ì´í„° ì°¾ê¸°
            for col in df.columns:
                col_lower = col.lower()
                if 'íšŒì‚¬' in col_lower or 'company' in col_lower:
                    for idx, val in df[col].items():
                        if isinstance(val, str) and len(val) > 1:
                            extracted_data['company_name'] = val
                            break
                
                if 'ìë³¸' in col_lower or 'capital' in col_lower or 'equity' in col_lower:
                    for idx, val in df[col].items():
                        if isinstance(val, (int, float)) and val > 0:
                            extracted_data['total_equity'] = int(val)
                            break
                
                if 'ë‹¹ê¸°ìˆœì´ìµ' in col_lower or 'profit' in col_lower or 'income' in col_lower:
                    profit_values = [val for idx, val in df[col].items() 
                                    if isinstance(val, (int, float)) and val > 0]
                    if len(profit_values) >= 3:
                        extracted_data['net_income1'] = int(profit_values[0])
                        extracted_data['net_income2'] = int(profit_values[1])
                        extracted_data['net_income3'] = int(profit_values[2])
                    elif len(profit_values) == 2:
                        extracted_data['net_income1'] = int(profit_values[0])
                        extracted_data['net_income2'] = int(profit_values[1])
                    elif len(profit_values) == 1:
                        extracted_data['net_income1'] = int(profit_values[0])
                
                if 'ì£¼ì‹' in col_lower or 'shares' in col_lower:
                    for idx, val in df[col].items():
                        if isinstance(val, (int, float)) and val > 0:
                            extracted_data['shares'] = int(val)
                            break
                
                if 'ì•¡ë©´' in col_lower or 'face' in col_lower:
                    for idx, val in df[col].items():
                        if isinstance(val, (int, float)) and val > 0:
                            extracted_data['share_price'] = int(val)
                            break
        except Exception as e:
            st.error(f"ì—‘ì…€ íŒŒì¼ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}")
    
    # Word íŒŒì¼ ì²˜ë¦¬
    elif file_ext in ['docx', 'doc'] and DOCX_AVAILABLE:
        try:
            doc = docx.Document(uploaded_file)
            paragraphs = [p.text for p in doc.paragraphs]
            text_content = "\n".join(paragraphs)
        except Exception as e:
            st.error(f"ì›Œë“œ íŒŒì¼ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}")
    
    # PDF íŒŒì¼ ì²˜ë¦¬
    elif file_ext == 'pdf' and PYPDF2_AVAILABLE:
        try:
            reader = PyPDF2.PdfReader(uploaded_file)
            text_content = ""
            for page in reader.pages:
                text_content += page.extract_text()
        except Exception as e:
            st.error(f"PDF íŒŒì¼ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}")
    
    # í…ìŠ¤íŠ¸ì—ì„œ íŒ¨í„´ìœ¼ë¡œ ë°ì´í„° ì¶”ì¶œ
    if text_content:
        # íšŒì‚¬ëª… íŒ¨í„´
        company_patterns = [
            r'íšŒì‚¬\s*ëª…\s*[:\s]+([^\n\r,]+)',
            r'íšŒì‚¬\s*[:\s]+([^\n\r,]+)',
            r'ìƒí˜¸\s*[:\s]+([^\n\r,]+)'
        ]
        for pattern in company_patterns:
            matches = re.findall(pattern, text_content)
            if matches:
                extracted_data['company_name'] = matches[0].strip()
                break
        
        # ìë³¸ì´ê³„ íŒ¨í„´
        capital_patterns = [
            r'ìë³¸ì´ê³„\s*[:\s]+([\d,\.]+)',
            r'ìë³¸\s*[:\s]+([\d,\.]+)',
            r'ì´ìë³¸\s*[:\s]+([\d,\.]+)',
            r'capital\s*[:\s]+([\d,\.]+)',
            r'equity\s*[:\s]+([\d,\.]+)'
        ]
        extracted_data['total_equity'] = find_with_patterns(text_content, capital_patterns)
        
        # ë‹¹ê¸°ìˆœì´ìµ íŒ¨í„´
        profit_patterns = [
            r'ë‹¹ê¸°ìˆœì´ìµ\s*[:\s]+([\d,\.]+)',
            r'ìˆœì´ìµ\s*[:\s]+([\d,\.]+)',
            r'profit\s*[:\s]+([\d,\.]+)',
            r'net income\s*[:\s]+([\d,\.]+)'
        ]
        profit_matches = []
        for pattern in profit_patterns:
            matches = re.findall(pattern, text_content, re.IGNORECASE)
            profit_matches.extend(matches)
        
        if len(profit_matches) >= 3:
            extracted_data['net_income1'] = int(re.sub(r'[^\d]', '', profit_matches[0]))
            extracted_data['net_income2'] = int(re.sub(r'[^\d]', '', profit_matches[1]))
            extracted_data['net_income3'] = int(re.sub(r'[^\d]', '', profit_matches[2]))
        elif len(profit_matches) == 2:
            extracted_data['net_income1'] = int(re.sub(r'[^\d]', '', profit_matches[0]))
            extracted_data['net_income2'] = int(re.sub(r'[^\d]', '', profit_matches[1]))
        elif len(profit_matches) == 1:
            extracted_data['net_income1'] = int(re.sub(r'[^\d]', '', profit_matches[0]))
        
        # ì£¼ì‹ìˆ˜ íŒ¨í„´
        shares_patterns = [
            r'ë°œí–‰ì£¼ì‹ìˆ˜\s*[:\s]+([\d,\.]+)',
            r'ì£¼ì‹ìˆ˜\s*[:\s]+([\d,\.]+)',
            r'shares\s*[:\s]+([\d,\.]+)'
        ]
        extracted_data['shares'] = find_with_patterns(text_content, shares_patterns)
        
        # ì•¡ë©´ê°€ íŒ¨í„´
        face_value_patterns = [
            r'ì•¡ë©´ê°€\s*[:\s]+([\d,\.]+)',
            r'ì•¡ë©´ê¸ˆì•¡\s*[:\s]+([\d,\.]+)',
            r'face value\s*[:\s]+([\d,\.]+)'
        ]
        extracted_data['share_price'] = find_with_patterns(text_content, face_value_patterns)
    
    return extracted_data

# ì¶”ì¶œí•œ ë°ì´í„°ë¥¼ ì„¸ì…˜ ìƒíƒœì— ì ìš©í•˜ëŠ” í•¨ìˆ˜
def apply_extracted_data(extracted_data):
    if 'company_name' in extracted_data and extracted_data['company_name']:
        st.session_state.company_name = extracted_data['company_name']
    if 'total_equity' in extracted_data and extracted_data['total_equity']:
        st.session_state.total_equity = extracted_data['total_equity']
    if 'net_income1' in extracted_data and extracted_data['net_income1']:
        st.session_state.net_income1 = extracted_data['net_income1']
    if 'net_income2' in extracted_data and extracted_data['net_income2']:
        st.session_state.net_income2 = extracted_data['net_income2']
    if 'net_income3' in extracted_data and extracted_data['net_income3']:
        st.session_state.net_income3 = extracted_data['net_income3']
    if 'shares' in extracted_data and extracted_data['shares']:
        st.session_state.shares = extracted_data['shares']
    if 'share_price' in extracted_data and extracted_data['share_price']:
        st.session_state.share_price = extracted_data['share_price']

# PDF ìƒì„± í•¨ìˆ˜
def generate_pdf():
    global FPDF_AVAILABLE  # global ë³€ìˆ˜ ì„ ì–¸ ì¶”ê°€
    
    if not FPDF_AVAILABLE:
        # FPDFë¥¼ ìë™ìœ¼ë¡œ ì„¤ì¹˜ ì‹œë„
        try:
            import subprocess
            subprocess.check_call(['pip', 'install', 'fpdf'])
            FPDF_AVAILABLE = True
        except:
            return None
    
    try:
        # PDF ê°ì²´ ìƒì„±
        pdf = FPDF()
        pdf.add_page()
        
        # ê¸°ë³¸ í°íŠ¸ ì‚¬ìš© (í•œê¸€ì€ í‘œì‹œë˜ì§€ ì•Šì„ ìˆ˜ ìˆìŒ)
        pdf.set_font('Arial', 'B', 16)
        
        # ì œëª©
        pdf.cell(190, 10, 'Unlisted Stock Valuation Report', 0, 1, 'C')
        pdf.ln(10)
        
        # ê¸°ë³¸ í°íŠ¸ ì„¤ì •
        pdf.set_font('Arial', '', 12)
        
        # í‰ê°€ ì •ë³´
        pdf.cell(190, 10, f'Date: {st.session_state.eval_date}', 0, 1)
        pdf.cell(190, 10, f'Company: {st.session_state.company_name}', 0, 1)
        pdf.ln(5)
        
        # ì¬ë¬´ ì •ë³´
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(190, 10, 'Financial Information', 0, 1)
        pdf.set_font('Arial', '', 12)
        pdf.cell(190, 10, f'Total Equity: {format_number(st.session_state.total_equity, True)} KRW', 0, 1)
        pdf.cell(190, 10, f'Net Income (Year 1): {format_number(st.session_state.net_income1, True)} KRW (Weight 3)', 0, 1)
        pdf.cell(190, 10, f'Net Income (Year 2): {format_number(st.session_state.net_income2, True)} KRW (Weight 2)', 0, 1)
        pdf.cell(190, 10, f'Net Income (Year 3): {format_number(st.session_state.net_income3, True)} KRW (Weight 1)', 0, 1)
        pdf.ln(5)
        
        # ì£¼ì‹ ì •ë³´
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(190, 10, 'Stock Information', 0, 1)
        pdf.set_font('Arial', '', 12)
        pdf.cell(190, 10, f'Total Shares: {format_number(st.session_state.shares)}', 0, 1)
        pdf.cell(190, 10, f'Face Value: {format_number(st.session_state.share_price)} KRW', 0, 1)
        pdf.cell(190, 10, f'Capitalization Rate: {st.session_state.interest_rate}%', 0, 1)
        pdf.ln(5)
        
        # í‰ê°€ ê²°ê³¼
        if st.session_state.stock_value:
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(190, 10, 'Valuation Results', 0, 1)
            pdf.set_font('Arial', '', 12)
            
            pdf.cell(190, 10, f'Net Asset Value per Share: {format_number(st.session_state.stock_value["netAssetPerShare"])} KRW', 0, 1)
            pdf.cell(190, 10, f'Income Value per Share: {format_number(st.session_state.stock_value["incomeValue"])} KRW', 0, 1)
            pdf.cell(190, 10, f'Asset Value with Goodwill: {format_number(st.session_state.stock_value["assetValueWithGoodwill"])} KRW', 0, 1)
            pdf.ln(5)
            
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(190, 10, f'Valuation Method: {st.session_state.stock_value["methodText"]}', 0, 1)
            pdf.cell(190, 10, f'Final Value per Share: {format_number(st.session_state.stock_value["finalValue"])} KRW', 0, 1)
            pdf.cell(190, 10, f'Total Company Value: {format_number(st.session_state.stock_value["totalValue"], True)} KRW', 0, 1)
        
        # PDFë¥¼ ë°”ì´íŠ¸ë¡œ ë³€í™˜ (ì•ˆì „í•œ ë°©ë²• ì‚¬ìš©)
        try:
            # ë°©ë²• 1: FPDF ê¸°ë³¸ ë°©ì‹
            pdf_output = pdf.output(dest='S').encode('latin-1')
            return pdf_output
        except Exception as e1:
            try:
                # ë°©ë²• 2: íŒŒì¼ì— ì €ì¥ í›„ ì½ê¸°
                import tempfile
                import os
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                    temp_path = tmp.name
                
                pdf.output(temp_path)
                with open(temp_path, 'rb') as f:
                    pdf_bytes = f.read()
                
                # ì„ì‹œ íŒŒì¼ ì‚­ì œ
                try:
                    os.unlink(temp_path)
                except:
                    pass
                
                return pdf_bytes
            except Exception as e2:
                return None
    
    except Exception as e:
        # ë””ë²„ê¹…ìš© ì˜¤ë¥˜ ì¶œë ¥
        import traceback
        print(f"PDF ìƒì„± ì˜¤ë¥˜: {e}")
        print(traceback.format_exc())
        return None

# HTML ë‹¤ìš´ë¡œë“œìš© ë‚´ìš© ìƒì„±
def create_html_content():
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>ë¹„ìƒì¥ì£¼ì‹ ê°€ì¹˜í‰ê°€ ë³´ê³ ì„œ</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #2c3e50; text-align: center; }}
            h2 {{ color: #3498db; margin-top: 20px; }}
            .info {{ margin-bottom: 5px; }}
            .result {{ margin-top: 10px; font-weight: bold; }}
        </style>
    </head>
    <body>
        <h1>ë¹„ìƒì¥ì£¼ì‹ ê°€ì¹˜í‰ê°€ ë³´ê³ ì„œ</h1>
        
        <h2>í‰ê°€ ì •ë³´</h2>
        <div class="info">í‰ê°€ ê¸°ì¤€ì¼: {st.session_state.eval_date}</div>
        <div class="info">íšŒì‚¬ëª…: {st.session_state.company_name}</div>
        
        <h2>ì¬ë¬´ ì •ë³´</h2>
        <div class="info">ìë³¸ì´ê³„: {format_number(st.session_state.total_equity)}ì›</div>
        <div class="info">1ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(st.session_state.net_income1)}ì› (ê°€ì¤‘ì¹˜ 3ë°°)</div>
        <div class="info">2ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(st.session_state.net_income2)}ì› (ê°€ì¤‘ì¹˜ 2ë°°)</div>
        <div class="info">3ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(st.session_state.net_income3)}ì› (ê°€ì¤‘ì¹˜ 1ë°°)</div>
        
        <h2>ì£¼ì‹ ì •ë³´</h2>
        <div class="info">ì´ ë°œí–‰ì£¼ì‹ìˆ˜: {format_number(st.session_state.shares)}ì£¼</div>
        <div class="info">ì•¡ë©´ê¸ˆì•¡: {format_number(st.session_state.share_price)}ì›</div>
        <div class="info">í™˜ì›ìœ¨: {st.session_state.interest_rate}%</div>
    """
    
    if st.session_state.stock_value:
        html_content += f"""
        <h2>í‰ê°€ ê²°ê³¼</h2>
        <div class="info">ìˆœìì‚°ê°€ì¹˜: {format_number(st.session_state.stock_value["netAssetPerShare"])}ì›</div>
        <div class="info">ìˆ˜ìµê°€ì¹˜: {format_number(st.session_state.stock_value["incomeValue"])}ì›</div>
        
        <div class="result">í‰ê°€ë°©ë²•: {st.session_state.stock_value["methodText"]}</div>
        <div class="result">ì£¼ë‹¹ í‰ê°€ì•¡: {format_number(st.session_state.stock_value["finalValue"])}ì›</div>
        <div class="result">ê¸°ì—… ì´ ê°€ì¹˜: {format_number(st.session_state.stock_value["totalValue"])}ì›</div>
        """
    
    html_content += """
    </body>
    </html>
    """
    return html_content

# CSV ë‹¤ìš´ë¡œë“œìš© ë‚´ìš© ìƒì„±
def create_csv_content():
    if not st.session_state.stock_value:
        return None
    
    # CSV ë°ì´í„° ìƒì„±
    data = {
        'í•­ëª©': ['í‰ê°€ ê¸°ì¤€ì¼', 'íšŒì‚¬ëª…', 'ìë³¸ì´ê³„', '1ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ', '2ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ', '3ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ',
                'ì´ ë°œí–‰ì£¼ì‹ìˆ˜', 'ì•¡ë©´ê¸ˆì•¡', 'í™˜ì›ìœ¨', 'ìˆœìì‚°ê°€ì¹˜', 'ìˆ˜ìµê°€ì¹˜', 
                'í‰ê°€ ë°©ë²•', 'ì£¼ë‹¹ í‰ê°€ì•¡', 'ê¸°ì—… ì´ ê°€ì¹˜'],
        'ê°’': [str(st.session_state.eval_date), st.session_state.company_name,
              st.session_state.total_equity, st.session_state.net_income1,
              st.session_state.net_income2, st.session_state.net_income3,
              st.session_state.shares, st.session_state.share_price,
              st.session_state.interest_rate, 
              st.session_state.stock_value["netAssetPerShare"],
              st.session_state.stock_value["incomeValue"],
              st.session_state.stock_value["methodText"],
              st.session_state.stock_value["finalValue"],
              st.session_state.stock_value["totalValue"]]
    }
    
    # DataFrame ìƒì„± í›„ CSVë¡œ ë³€í™˜
    df = pd.DataFrame(data)
    csv = df.to_csv(index=False).encode('utf-8')
    return csv

# ë¹„ìƒì¥ì£¼ì‹ ê°€ì¹˜ ê³„ì‚° í•¨ìˆ˜
def calculate_stock_value():
    # ì…ë ¥ê°’ ê°€ì ¸ì˜¤ê¸°
    total_equity = st.session_state.total_equity
    net_income1 = st.session_state.net_income1
    net_income2 = st.session_state.net_income2
    net_income3 = st.session_state.net_income3
    shares = st.session_state.shares
    owned_shares = st.session_state.owned_shares
    interest_rate = st.session_state.interest_rate
    evaluation_method = st.session_state.evaluation_method
    eval_date = st.session_state.eval_date
    
    # ê³„ì‚° ë¡œì§
    net_asset_per_share = total_equity / shares
    weighted_income = (net_income1 * 3 + net_income2 * 2 + net_income3 * 1) / 6
    weighted_income_per_share = weighted_income / shares
    weighted_income_per_share_50 = weighted_income_per_share * 0.5
    equity_return = (total_equity * (interest_rate / 100)) / shares
    annuity_factor = 3.7908
    goodwill = max(0, (weighted_income_per_share_50 - equity_return) * annuity_factor)
    asset_value_with_goodwill = net_asset_per_share + goodwill
    income_value = weighted_income_per_share * (100 / interest_rate)
    
    if evaluation_method == "ë¶€ë™ì‚° ê³¼ë‹¤ë²•ì¸":
        stock_value = (asset_value_with_goodwill * 0.6) + (income_value * 0.4)
        net_asset_80_percent = net_asset_per_share * 0.8
        final_value = max(stock_value, net_asset_80_percent)
        method_text = 'ë¶€ë™ì‚° ê³¼ë‹¤ë²•ì¸: (ìì‚°ê°€ì¹˜Ã—0.6 + ìˆ˜ìµê°€ì¹˜Ã—0.4)'
    elif evaluation_method == "ìˆœìì‚°ê°€ì¹˜ë§Œ í‰ê°€":
        final_value = net_asset_per_share
        method_text = 'ìˆœìì‚°ê°€ì¹˜ë§Œ í‰ê°€'
    else:
        stock_value = (income_value * 0.6) + (asset_value_with_goodwill * 0.4)
        net_asset_80_percent = net_asset_per_share * 0.8
        final_value = max(stock_value, net_asset_80_percent)
        method_text = 'ì¼ë°˜ë²•ì¸: (ìˆ˜ìµê°€ì¹˜Ã—0.6 + ìì‚°ê°€ì¹˜Ã—0.4)'
    
    total_value = final_value * shares
    owned_value = final_value * owned_shares
    increase_percentage = round((final_value / net_asset_per_share) * 100) if net_asset_per_share > 0 else 0
    
    return {
        "evalDate": eval_date,
        "netAssetPerShare": net_asset_per_share,
        "assetValueWithGoodwill": asset_value_with_goodwill,
        "incomeValue": income_value,
        "finalValue": final_value,
        "totalValue": total_value,
        "ownedValue": owned_value,
        "methodText": method_text,
        "increasePercentage": increase_percentage,
        "weightedIncome": weighted_income
    }

# í˜ì´ì§€ í—¤ë”
st.title("ë¹„ìƒì¥ì£¼ì‹ ê°€ì¹˜í‰ê°€")

# íŒŒì¼ ì—…ë¡œë“œ ì„¹ì…˜ (expanderë¡œ ìˆ¨ê¹€)
with st.expander("ğŸ“¤ íŒŒì¼ ì—…ë¡œë“œ (ì—‘ì…€, PDF, ì›Œë“œ)", expanded=False):
    st.markdown("<div class='upload-section'>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader("ì¬ë¬´ì •ë³´ê°€ í¬í•¨ëœ íŒŒì¼ì„ ì—…ë¡œë“œí•˜ì„¸ìš”", 
                                    type=['xlsx', 'xls', 'pdf', 'docx', 'doc'])
    
    if uploaded_file is not None:
        st.write(f"íŒŒì¼ëª…: {uploaded_file.name}, íŒŒì¼í¬ê¸°: {uploaded_file.size} bytes")
        
        # ë°ì´í„° ì¶”ì¶œ ë²„íŠ¼
        if st.button("ğŸ“¥ ë°ì´í„° ì¶”ì¶œ", key="extract_data"):
            with st.spinner('íŒŒì¼ì—ì„œ ë°ì´í„°ë¥¼ ì¶”ì¶œ ì¤‘ì…ë‹ˆë‹¤...'):
                extracted_data = extract_data_from_file(uploaded_file)
                
                # ì¶”ì¶œëœ ë°ì´í„° í‘œì‹œ
                if extracted_data:
                    st.success("ë‹¤ìŒ ë°ì´í„°ê°€ ì¶”ì¶œë˜ì—ˆìŠµë‹ˆë‹¤:")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if 'company_name' in extracted_data:
                            st.write(f"íšŒì‚¬ëª…: {extracted_data['company_name']}")
                        if 'total_equity' in extracted_data:
                            st.write(f"ìë³¸ì´ê³„: {format_number(extracted_data['total_equity'])}ì›")
                        if 'shares' in extracted_data:
                            st.write(f"ì´ ë°œí–‰ì£¼ì‹ìˆ˜: {format_number(extracted_data['shares'])}ì£¼")
                    
                    with col2:
                        if 'net_income1' in extracted_data:
                            st.write(f"1ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(extracted_data['net_income1'])}ì›")
                        if 'net_income2' in extracted_data:
                            st.write(f"2ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(extracted_data['net_income2'])}ì›")
                        if 'net_income3' in extracted_data:
                            st.write(f"3ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµ: {format_number(extracted_data['net_income3'])}ì›")
                        if 'share_price' in extracted_data:
                            st.write(f"ì•¡ë©´ê¸ˆì•¡: {format_number(extracted_data['share_price'])}ì›")
                    
                    # ë°ì´í„° ì ìš© ë²„íŠ¼
                    if st.button("ğŸ’¾ ì¶”ì¶œëœ ë°ì´í„° ì ìš©", key="apply_data"):
                        apply_extracted_data(extracted_data)
                        st.success("ë°ì´í„°ê°€ ì ìš©ë˜ì—ˆìŠµë‹ˆë‹¤.")
                        st.experimental_rerun()
                else:
                    st.warning("íŒŒì¼ì—ì„œ ì¶”ì¶œ ê°€ëŠ¥í•œ ë°ì´í„°ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")
    st.markdown("</div>", unsafe_allow_html=True)

# í‰ê°€ ê¸°ì¤€ì¼ ì„¤ì •
with st.expander("í‰ê°€ ê¸°ì¤€ì¼", expanded=True):
    col1, col2 = st.columns([1, 2])
    
    with col1:
        eval_date = st.date_input(
            "í‰ê°€ ê¸°ì¤€ì¼",
            value=st.session_state.eval_date,
            help="ë¹„ìƒì¥ì£¼ì‹ í‰ê°€ì˜ ê¸°ì¤€ì´ ë˜ëŠ” ë‚ ì§œì…ë‹ˆë‹¤. ë³´í†µ ê²°ì‚°ì¼ì´ë‚˜ í‰ê°€ê°€ í•„ìš”í•œ ì‹œì ìœ¼ë¡œ ì„¤ì •í•©ë‹ˆë‹¤.",
            key="eval_date_input"
        )
    
    with col2:
        st.markdown("<div class='field-description'>í‰ê°€ ê¸°ì¤€ì¼ì€ ìë³¸ì´ê³„, ë‹¹ê¸°ìˆœì´ìµ ë“± ì¬ë¬´ì •ë³´ì˜ ê¸°ì¤€ ì‹œì ì…ë‹ˆë‹¤. ì¼ë°˜ì ìœ¼ë¡œ ê°€ì¥ ìµœê·¼ ê²°ì‚°ì¼ì„ ì‚¬ìš©í•©ë‹ˆë‹¤.</div>", unsafe_allow_html=True)

# íšŒì‚¬ ì •ë³´ ì…ë ¥
with st.expander("íšŒì‚¬ ì •ë³´", expanded=True):
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("<div class='section-header'>íšŒì‚¬ëª…</div>", unsafe_allow_html=True)
        company_name = st.text_input(
            "íšŒì‚¬ëª…", 
            value=st.session_state.company_name,
            help="í‰ê°€ ëŒ€ìƒ íšŒì‚¬ì˜ ì •ì‹ ëª…ì¹­ì„ ì…ë ¥í•˜ì„¸ìš”.",
            key="company_name_input",
            label_visibility="collapsed"
        )
    
    with col2:
        st.markdown("<div class='section-header'>ìë³¸ì´ê³„ (ì›)</div>", unsafe_allow_html=True)
        
        # ë‹¨ìœ„ ì„ íƒ ì¶”ê°€
        total_equity_unit = st.radio(
            "ë‹¨ìœ„ ì„ íƒ",
            options=["ì›", "ì²œì›"],
            horizontal=True,
            key="total_equity_unit_radio",
            label_visibility="collapsed",
            index=0 if st.session_state.total_equity_unit == "ì›" else 1
        )
        st.session_state.total_equity_unit = total_equity_unit
        
        # ìˆ«ì ì…ë ¥ ì»¨í…Œì´ë„ˆ
        st.markdown("<div class='number-input-container'>", unsafe_allow_html=True)
        
        # ì…ë ¥ê°’ ë³€í™˜ (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        display_value = st.session_state.total_equity
        if total_equity_unit == "ì²œì›":
            display_value = display_value // 1000
            
        total_equity = st.number_input(
            "ìë³¸ì´ê³„", 
            value=display_value, 
            min_value=0, 
            format="%d",
            help="í‰ê°€ ê¸°ì¤€ì¼ í˜„ì¬ íšŒì‚¬ì˜ ëŒ€ì°¨ëŒ€ì¡°í‘œìƒ ìë³¸ì´ê³„ë¥¼ ì…ë ¥í•˜ì„¸ìš”.",
            key="total_equity_input",
            label_visibility="collapsed"
        )
        
        # ë¦¬ì…‹ ë²„íŠ¼ ì¶”ê°€
        reset_btn = st.button("âœ–", key="reset_total_equity", help="ì…ë ¥ê°’ ì´ˆê¸°í™”")
        if reset_btn:
            if total_equity_unit == "ì›":
                st.session_state.total_equity_input = 0
            else:
                st.session_state.total_equity_input = 0
            st.experimental_rerun()
            
        st.markdown("</div>", unsafe_allow_html=True)
        
        # ì‹¤ì œ ê°’ ê³„ì‚° (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        actual_value = total_equity
        if total_equity_unit == "ì²œì›":
            actual_value = total_equity * 1000
            
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.total_equity = actual_value
        
        # ê¸ˆì•¡ í‘œì‹œ
        if total_equity_unit == "ì›":
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value)}ì›</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value // 1000)}ì²œì›</div>", unsafe_allow_html=True)
        st.markdown("<div class='field-description'>ì¬ë¬´ìƒíƒœí‘œ(ëŒ€ì°¨ëŒ€ì¡°í‘œ)ìƒì˜ ìë³¸ì´ê³„ ê¸ˆì•¡ì…ë‹ˆë‹¤. í‰ê°€ê¸°ì¤€ì¼ í˜„ì¬ì˜ ê¸ˆì•¡ì„ ì…ë ¥í•˜ì„¸ìš”.</div>", unsafe_allow_html=True)

# ë‹¹ê¸°ìˆœì´ìµ ì…ë ¥
with st.expander("ë‹¹ê¸°ìˆœì´ìµ (ìµœê·¼ 3ê°œë…„)", expanded=True):
    st.markdown("<div class='field-description'>ìµœê·¼ 3ê°œ ì‚¬ì—…ì—°ë„ì˜ ë‹¹ê¸°ìˆœì´ìµì„ ì…ë ¥í•˜ì„¸ìš”. ê° ì—°ë„ë³„ë¡œ ê°€ì¤‘ì¹˜ê°€ ë‹¤ë¥´ê²Œ ì ìš©ë©ë‹ˆë‹¤.</div>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("##### 1ë…„ ì „ (ê°€ì¤‘ì¹˜ 3ë°°)")
        
        # ë‹¨ìœ„ ì„ íƒ ì¶”ê°€
        net_income1_unit = st.radio(
            "ë‹¨ìœ„ ì„ íƒ (1ë…„ ì „)",
            options=["ì›", "ì²œì›"],
            horizontal=True,
            key="net_income1_unit_radio",
            label_visibility="collapsed",
            index=0 if st.session_state.net_income1_unit == "ì›" else 1
        )
        st.session_state.net_income1_unit = net_income1_unit
        
        # ìˆ«ì ì…ë ¥ ì»¨í…Œì´ë„ˆ
        st.markdown("<div class='number-input-container'>", unsafe_allow_html=True)
        
        # ì…ë ¥ê°’ ë³€í™˜ (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        display_value = st.session_state.net_income1
        if net_income1_unit == "ì²œì›":
            display_value = display_value // 1000
            
        net_income1 = st.number_input(
            "ë‹¹ê¸°ìˆœì´ìµ (ì›)", 
            value=display_value, 
            format="%d",
            help="ê°€ì¥ ìµœê·¼ ì—°ë„ì˜ ë‹¹ê¸°ìˆœì´ìµì…ë‹ˆë‹¤. 3ë°° ê°€ì¤‘ì¹˜ê°€ ì ìš©ë©ë‹ˆë‹¤.",
            key="income_year1_input",
            label_visibility="collapsed"
        )
        
        # ë¦¬ì…‹ ë²„íŠ¼ ì¶”ê°€
        reset_btn = st.button("âœ–", key="reset_net_income1", help="ì…ë ¥ê°’ ì´ˆê¸°í™”")
        if reset_btn:
            st.session_state.income_year1_input = 0
            st.experimental_rerun()
            
        st.markdown("</div>", unsafe_allow_html=True)
        
        # ì‹¤ì œ ê°’ ê³„ì‚° (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        actual_value = net_income1
        if net_income1_unit == "ì²œì›":
            actual_value = net_income1 * 1000
            
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.net_income1 = actual_value
        
        # ê¸ˆì•¡ í‘œì‹œ
        if net_income1_unit == "ì›":
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value)}ì›</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value // 1000)}ì²œì›</div>", unsafe_allow_html=True)
        
    with col2:
        st.markdown("##### 2ë…„ ì „ (ê°€ì¤‘ì¹˜ 2ë°°)")
        
        # ë‹¨ìœ„ ì„ íƒ ì¶”ê°€
        net_income2_unit = st.radio(
            "ë‹¨ìœ„ ì„ íƒ (2ë…„ ì „)",
            options=["ì›", "ì²œì›"],
            horizontal=True,
            key="net_income2_unit_radio",
            label_visibility="collapsed",
            index=0 if st.session_state.net_income2_unit == "ì›" else 1
        )
        st.session_state.net_income2_unit = net_income2_unit
        
        # ìˆ«ì ì…ë ¥ ì»¨í…Œì´ë„ˆ
        st.markdown("<div class='number-input-container'>", unsafe_allow_html=True)
        
        # ì…ë ¥ê°’ ë³€í™˜ (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        display_value = st.session_state.net_income2
        if net_income2_unit == "ì²œì›":
            display_value = display_value // 1000
            
        net_income2 = st.number_input(
            "ë‹¹ê¸°ìˆœì´ìµ (ì›)", 
            value=display_value, 
            format="%d",
            help="2ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµì…ë‹ˆë‹¤. 2ë°° ê°€ì¤‘ì¹˜ê°€ ì ìš©ë©ë‹ˆë‹¤.",
            key="income_year2_input",
            label_visibility="collapsed"
        )
        
        # ë¦¬ì…‹ ë²„íŠ¼ ì¶”ê°€
        reset_btn = st.button("âœ–", key="reset_net_income2", help="ì…ë ¥ê°’ ì´ˆê¸°í™”")
        if reset_btn:
            st.session_state.income_year2_input = 0
            st.experimental_rerun()
            
        st.markdown("</div>", unsafe_allow_html=True)
        
        # ì‹¤ì œ ê°’ ê³„ì‚° (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        actual_value = net_income2
        if net_income2_unit == "ì²œì›":
            actual_value = net_income2 * 1000
            
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.net_income2 = actual_value
        
        # ê¸ˆì•¡ í‘œì‹œ
        if net_income2_unit == "ì›":
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value)}ì›</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value // 1000)}ì²œì›</div>", unsafe_allow_html=True)
        
    with col3:
        st.markdown("##### 3ë…„ ì „ (ê°€ì¤‘ì¹˜ 1ë°°)")
        
        # ë‹¨ìœ„ ì„ íƒ ì¶”ê°€
        net_income3_unit = st.radio(
            "ë‹¨ìœ„ ì„ íƒ (3ë…„ ì „)",
            options=["ì›", "ì²œì›"],
            horizontal=True,
            key="net_income3_unit_radio",
            label_visibility="collapsed",
            index=0 if st.session_state.net_income3_unit == "ì›" else 1
        )
        st.session_state.net_income3_unit = net_income3_unit
        
        # ìˆ«ì ì…ë ¥ ì»¨í…Œì´ë„ˆ
        st.markdown("<div class='number-input-container'>", unsafe_allow_html=True)
        
        # ì…ë ¥ê°’ ë³€í™˜ (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        display_value = st.session_state.net_income3
        if net_income3_unit == "ì²œì›":
            display_value = display_value // 1000
            
        net_income3 = st.number_input(
            "ë‹¹ê¸°ìˆœì´ìµ (ì›)", 
            value=display_value, 
            format="%d",
            help="3ë…„ ì „ ë‹¹ê¸°ìˆœì´ìµì…ë‹ˆë‹¤. 1ë°° ê°€ì¤‘ì¹˜ê°€ ì ìš©ë©ë‹ˆë‹¤.",
            key="income_year3_input",
            label_visibility="collapsed"
        )
        
        # ë¦¬ì…‹ ë²„íŠ¼ ì¶”ê°€
        reset_btn = st.button("âœ–", key="reset_net_income3", help="ì…ë ¥ê°’ ì´ˆê¸°í™”")
        if reset_btn:
            st.session_state.income_year3_input = 0
            st.experimental_rerun()
            
        st.markdown("</div>", unsafe_allow_html=True)
        
        # ì‹¤ì œ ê°’ ê³„ì‚° (ì²œì› ë‹¨ìœ„ë¡œ ì…ë ¥í–ˆë‹¤ë©´ ì› ë‹¨ìœ„ë¡œ ë³€í™˜)
        actual_value = net_income3
        if net_income3_unit == "ì²œì›":
            actual_value = net_income3 * 1000
            
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.net_income3 = actual_value
        
        # ê¸ˆì•¡ í‘œì‹œ
        if net_income3_unit == "ì›":
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value)}ì›</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(actual_value // 1000)}ì²œì›</div>", unsafe_allow_html=True)

# ì£¼ì‹ ì •ë³´ ì…ë ¥
with st.expander("ì£¼ì‹ ì •ë³´", expanded=True):
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col1:
        st.markdown("<div class='section-header'>ì´ ë°œí–‰ì£¼ì‹ìˆ˜</div>", unsafe_allow_html=True)
        
        shares = st.number_input(
            "ì´ ë°œí–‰ì£¼ì‹ìˆ˜", 
            value=st.session_state.shares, 
            min_value=1, 
            format="%d",
            help="íšŒì‚¬ê°€ ë°œí–‰í•œ ì´ ì£¼ì‹ìˆ˜ì…ë‹ˆë‹¤.",
            key="shares_input",
            label_visibility="collapsed"
        )
        
        st.markdown(f"<div class='amount-display'>ì´ {format_number(shares)}ì£¼</div>", unsafe_allow_html=True)
        
    with col2:
        st.markdown("<div class='section-header'>ì•¡ë©´ê¸ˆì•¡ (ì›)</div>", unsafe_allow_html=True)
        
        share_price = st.number_input(
            "ì•¡ë©´ê¸ˆì•¡ (ì›)", 
            value=st.session_state.share_price, 
            min_value=0, 
            format="%d",
            help="ì£¼ì‹ 1ì£¼ë‹¹ ì•¡ë©´ê°€ì•¡ì…ë‹ˆë‹¤. ì¼ë°˜ì ìœ¼ë¡œ 100ì›, 500ì›, 1,000ì›, 5,000ì› ë“±ìœ¼ë¡œ ì„¤ì •ë©ë‹ˆë‹¤.",
            key="share_price_input",
            label_visibility="collapsed"
        )
        
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.share_price = share_price
        
        # ê¸ˆì•¡ í‘œì‹œ
        st.markdown(f"<div class='amount-display'>ê¸ˆì•¡: {format_number(share_price)}ì›</div>", unsafe_allow_html=True)
        
    with col3:
        st.markdown("<div class='section-header'>í™˜ì›ìœ¨</div>", unsafe_allow_html=True)
        # í™˜ì›ìœ¨ 10%ë¡œ ê³ ì • í‘œì‹œ (ìŠ¬ë¼ì´ë” ì œê±°)
        st.markdown("<div style='font-size:14px; color:#666;'>í™˜ì›ìœ¨ì€ 10%ë¡œ ê³ ì •ë˜ì–´ ìˆìŠµë‹ˆë‹¤.</div>", unsafe_allow_html=True)
        # ì„¸ì…˜ ìƒíƒœ ë³€ìˆ˜ ì—…ë°ì´íŠ¸
        st.session_state.interest_rate = 10

# ì£¼ì£¼ ì •ë³´ ì…ë ¥
with st.expander("ì£¼ì£¼ ì •ë³´", expanded=True):
    col1, col2 = st.columns([2, 2])
    
    with col1:
        shareholder_count = st.selectbox(
            "ì…ë ¥í•  ì£¼ì£¼ ìˆ˜",
            options=[1, 2, 3, 4, 5],
            index=st.session_state.shareholder_count - 1,
            key="shareholder_count_select",
            help="ì…ë ¥í•  ì£¼ì£¼ì˜ ìˆ˜ë¥¼ ì„ íƒí•˜ì„¸ìš”. ìµœëŒ€ 5ëª…ê¹Œì§€ ì…ë ¥ ê°€ëŠ¥í•©ë‹ˆë‹¤."
        )
    
    st.session_state.shareholder_count = shareholder_count
    
    st.markdown("<div style='margin:10px 0; padding:10px; background-color:#f0f7fb; border-radius:5px; font-weight:bold;'>"
                f"ì£¼ì£¼ ì •ë³´ë¥¼ ì…ë ¥í•˜ì„¸ìš” (ì„ íƒëœ ì£¼ì£¼ ìˆ˜: {shareholder_count}ëª…)"
                "</div>", unsafe_allow_html=True)
    
    st.markdown("<div class='field-description'>íšŒì‚¬ì˜ ì£¼ì£¼ ì •ë³´ë¥¼ ì…ë ¥í•˜ì„¸ìš”. ì£¼ì£¼ë³„ ë³´ìœ  ì£¼ì‹ìˆ˜ëŠ” ë°œí–‰ì£¼ì‹ ì´ìˆ˜ë¥¼ ì´ˆê³¼í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.</div>", unsafe_allow_html=True)
    
    # ì´ ì£¼ì‹ìˆ˜ í™•ì¸ì„ ìœ„í•œ ë³€ìˆ˜
    total_owned_shares = 0
    
    # ì„ íƒí•œ ìˆ˜ë§Œí¼ì˜ ì£¼ì£¼ ì •ë³´ ì…ë ¥
    shareholders = []
    for i in range(shareholder_count):
        col1, col2 = st.columns([1, 1])
        with col1:
            name = st.text_input(
                f"ì£¼ì£¼ {i+1} ì´ë¦„", 
                value=st.session_state.shareholders[i]["name"] if i < len(st.session_state.shareholders) else "",
                key=f"shareholder_name_input_{i}",
                help=f"ì£¼ì£¼ {i+1}ì˜ ì´ë¦„ì„ ì…ë ¥í•˜ì„¸ìš”."
            )
        
        with col2:
            # ìˆ«ì ì…ë ¥ ì»¨í…Œì´ë„ˆ
            st.markdown("<div class='number-input-container'>", unsafe_allow_html=True)
            
            shares_owned = st.number_input(
                f"ë³´ìœ  ì£¼ì‹ìˆ˜", 
                value=st.session_state.shareholders[i]["shares"] if i < len(st.session_state.shareholders) else 0, 
                min_value=0,
                max_value=shares,
                format="%d",
                key=f"shareholder_shares_input_{i}",
                help=f"ì£¼ì£¼ {i+1}ì˜ ë³´ìœ  ì£¼ì‹ìˆ˜ë¥¼ ì…ë ¥í•˜ì„¸ìš”."
            )
            
            # ë¦¬ì…‹ ë²„íŠ¼ ì¶”ê°€
            reset_btn = st.button("âœ–", key=f"reset_shareholder_shares_{i}", help="ì…ë ¥ê°’ ì´ˆê¸°í™”")
            if reset_btn:
                st.session_state[f"shareholder_shares_input_{i}"] = 0
                st.experimental_rerun()
                
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown(f"<div class='amount-display'>{format_number(shares_owned)}ì£¼</div>", unsafe_allow_html=True)
        
        # êµ¬ë¶„ì„  ì¶”ê°€
        if i < shareholder_count - 1:
            st.markdown("<hr style='margin:10px 0; border-color:#eee;'>", unsafe_allow_html=True)
        
        shareholders.append({"name": name, "shares": shares_owned})
        total_owned_shares += shares_owned
    
    # ë‚˜ë¨¸ì§€ ì£¼ì£¼ ì •ë³´ ë³´ì¡´
    for i in range(shareholder_count, 5):
        if i < len(st.session_state.shareholders):
            shareholders.append(st.session_state.shareholders[i])
        else:
            shareholders.append({"name": "", "shares": 0})
    
    # ì…ë ¥ëœ ì£¼ì‹ìˆ˜ í•©ê³„ í™•ì¸
    ownership_percent = round(total_owned_shares/shares*100, 2) if shares > 0 else 0
    
    st.markdown(f"""
    <div style='margin-top:15px; padding:10px; border-radius:5px; 
         background-color:{"#e6f3e6" if total_owned_shares <= shares else "#f8d7da"}; 
         color:{"#0c5460" if total_owned_shares <= shares else "#721c24"};
         font-weight:bold;'>
        {'âœ…' if total_owned_shares <= shares else 'âš ï¸'} 
        ì£¼ì£¼ë“¤ì˜ ì´ ë³´ìœ  ì£¼ì‹ìˆ˜: {format_number(total_owned_shares)}ì£¼ 
        (ë°œí–‰ì£¼ì‹ìˆ˜ì˜ {ownership_percent}%)
        {' â€» ë°œí–‰ì£¼ì‹ìˆ˜ë¥¼ ì´ˆê³¼í–ˆìŠµë‹ˆë‹¤.' if total_owned_shares > shares else ''}
    </div>
    """, unsafe_allow_html=True)
    
    # ëŒ€í‘œì´ì‚¬ ë³´ìœ  ì£¼ì‹ìˆ˜ ì„¤ì •
    owned_shares = shareholders[0]["shares"] if shareholders and shareholders[0]["name"] else 0

# í‰ê°€ ë°©ì‹ ì„ íƒ
with st.expander("í‰ê°€ ë°©ì‹ ì„ íƒ", expanded=True):
    st.markdown("<div class='section-header'>ë¹„ìƒì¥ì£¼ì‹ í‰ê°€ ë°©ë²•</div>", unsafe_allow_html=True)
    evaluation_method = st.selectbox(
        "ë¹„ìƒì¥ì£¼ì‹ í‰ê°€ ë°©ë²•",
        ("ì¼ë°˜ë²•ì¸", "ë¶€ë™ì‚° ê³¼ë‹¤ë²•ì¸", "ìˆœìì‚°ê°€ì¹˜ë§Œ í‰ê°€"),
        index=["ì¼ë°˜ë²•ì¸", "ë¶€ë™ì‚° ê³¼ë‹¤ë²•ì¸", "ìˆœìì‚°ê°€ì¹˜ë§Œ í‰ê°€"].index(st.session_state.evaluation_method),
        key="evaluation_method_select",
        help="ìƒì†ì„¸ ë° ì¦ì—¬ì„¸ë²• ì‹œí–‰ë ¹ ì œ54ì¡°ì— ê·¼ê±°í•œ í‰ê°€ ë°©ë²•ì„ ì„ íƒí•˜ì„¸ìš”.",
        label_visibility="collapsed"
    )
    
    st.markdown("""
    <div style='background-color:#f0f7fb; padding:15px; border-radius:5px; margin-top:15px;'>
    <span style='font-weight:bold; font-size:16px;'>ğŸ“Œ í‰ê°€ë°©ì‹ ì„¤ëª…</span>
    <ul style='margin-top:10px; margin-bottom:0; padding-left:20px;'>
        <li><strong>ì¼ë°˜ë²•ì¸</strong>: ëŒ€ë¶€ë¶„ì˜ ë²•ì¸ì— ì ìš© (ìˆ˜ìµê°€ì¹˜ 60% + ìì‚°ê°€ì¹˜ 40%)</li>
        <li><strong>ë¶€ë™ì‚° ê³¼ë‹¤ë²•ì¸</strong>: ë¶€ë™ì‚°ì´ ìì‚°ì˜ 50% ì´ìƒì¸ ë²•ì¸ (ìì‚°ê°€ì¹˜ 60% + ìˆ˜ìµê°€ì¹˜ 40%)</li>
        <li><strong>ìˆœìì‚°ê°€ì¹˜ë§Œ í‰ê°€</strong>: íŠ¹ìˆ˜í•œ ê²½ìš° (ì„¤ë¦½ 1ë…„ ë¯¸ë§Œ ë“±) (ìˆœìì‚°ê°€ì¹˜ 100%)</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<div class='field-description'>ìƒì†ì„¸ ë° ì¦ì—¬ì„¸ë²• ì‹œí–‰ë ¹ ì œ54ì¡°ì— ê·¼ê±°í•œ í‰ê°€ë°©ë²•ì…ë‹ˆë‹¤. íšŒì‚¬ì˜ íŠ¹ì„±ì— ë§ëŠ” ë°©ë²•ì„ ì„ íƒí•˜ì„¸ìš”.</div>", unsafe_allow_html=True)

# ê³„ì‚° ë²„íŠ¼
st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

if st.button("ë¹„ìƒì¥ì£¼ì‹ í‰ê°€í•˜ê¸°", type="primary", use_container_width=True, key="evaluate_button"):
    with st.spinner("ê³„ì‚° ì¤‘..."):
        # ì„¸ì…˜ ìƒíƒœ ì—…ë°ì´íŠ¸
        st.session_state.eval_date = eval_date
        st.session_state.company_name = company_name
        st.session_state.total_equity = st.session_state.total_equity
        st.session_state.net_income1 = st.session_state.net_income1
        st.session_state.net_income2 = st.session_state.net_income2
        st.session_state.net_income3 = st.session_state.net_income3
        st.session_state.shares = shares
        st.session_state.owned_shares = owned_shares
        st.session_state.share_price = st.session_state.share_price
        st.session_state.interest_rate = 10  # í™˜ì›ìœ¨ 10% ê³ ì •
        st.session_state.evaluation_method = evaluation_method
        st.session_state.shareholders = shareholders
        
        # ì£¼ì‹ ê°€ì¹˜ ê³„ì‚°
        st.session_state.stock_value = calculate_stock_value()
        st.session_state.evaluated = True
        
        st.success(f"âœ… ê³„ì‚°ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤. í‰ê°€ê¸°ì¤€ì¼: {eval_date.strftime('%Yë…„ %mì›” %dì¼')} ê¸°ì¤€")
        st.balloons()

# ê²°ê³¼ í‘œì‹œ
if st.session_state.evaluated and st.session_state.stock_value:
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader("í‰ê°€ ê²°ê³¼")
    
    # ê²°ê³¼ ë°ì´í„°
    stock_value = st.session_state.stock_value
    
    # 3ì—´ ë ˆì´ì•„ì›ƒ
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("##### ìˆœìì‚°ê°€ì¹˜")
        st.markdown(f"<div style='font-size:20px; font-weight:bold;'>{format_number(stock_value['netAssetPerShare'])}ì›/ì£¼</div>", unsafe_allow_html=True)
        st.markdown(f"ì´ {format_number(stock_value['netAssetPerShare'] * shares)}ì›")
    
    with col2:
        st.markdown("##### ìˆ˜ìµê°€ì¹˜")
        st.markdown(f"<div style='font-size:20px; font-weight:bold;'>{format_number(stock_value['incomeValue'])}ì›/ì£¼</div>", unsafe_allow_html=True)
        st.markdown(f"ì´ {format_number(stock_value['incomeValue'] * shares)}ì›")
    
    with col3:
        st.markdown("##### ìµœì¢… í‰ê°€ì•¡")
        st.markdown(f"<div style='font-size:24px; font-weight:bold; color:#0066cc;'>{format_number(stock_value['finalValue'])}ì›/ì£¼</div>", unsafe_allow_html=True)
        st.markdown(f"ì´ {format_number(stock_value['totalValue'])}ì›")
    
    # í‰ê°€ ë°©ë²• ì„¤ëª…
    st.markdown(f"""
    <div style='margin-top:20px; padding:15px; background-color:#f8f9fa; border-radius:8px;'>
    <h5>ì ìš©ëœ í‰ê°€ ë°©ì‹: {stock_value['methodText']}</h5>
    <p>ìµœì¢… í‰ê°€ì•¡ì€ ìë³¸ì´ê³„ {format_number(total_equity)}ì›ì„ ê¸°ì¤€ìœ¼ë¡œ ê³„ì‚°ë˜ì—ˆìœ¼ë©°, 
    ê°€ì¤‘í‰ê·  ë‹¹ê¸°ìˆœì´ìµ {format_number(stock_value['weightedIncome'])}ì›ê³¼ 
    í™˜ì›ìœ¨ {st.session_state.interest_rate}%ë¥¼ ì ìš©í•˜ì—¬ ì‚°ì¶œë˜ì—ˆìŠµë‹ˆë‹¤.</p>
    </div>
    """, unsafe_allow_html=True)

    # ë‹¤ìš´ë¡œë“œ ì„¹ì…˜
    with st.expander("ğŸ“¥ í‰ê°€ ê²°ê³¼ ë‹¤ìš´ë¡œë“œ", expanded=False):
        st.markdown("<div class='download-section'>", unsafe_allow_html=True)
        tab1, tab2, tab3 = st.tabs(["PDF", "HTML", "CSV"])
        
        with tab1:
            # PDF ë‹¤ìš´ë¡œë“œ ë²„íŠ¼ ì§ì ‘ í‘œì‹œí•˜ê¸°
            if st.button("PDF ìƒì„±í•˜ê¸°", key="generate_pdf", type="primary"):
                with st.spinner("PDF ìƒì„± ì¤‘..."):
                    pdf_data = generate_pdf()
                    if pdf_data:
                        st.success("PDF ìƒì„± ì™„ë£Œ! ì•„ë˜ ë²„íŠ¼ì„ í´ë¦­í•˜ì—¬ ë‹¤ìš´ë¡œë“œí•˜ì„¸ìš”.")
                        st.download_button(
                            label="ğŸ“„ PDF íŒŒì¼ ë‹¤ìš´ë¡œë“œ",
                            data=pdf_data,
                            file_name=f"ë¹„ìƒì¥ì£¼ì‹_í‰ê°€_{st.session_state.company_name}_{st.session_state.eval_date}.pdf",
                            mime="application/pdf"
                        )
                    else:
                        # ëŒ€ì²´ PDF ìƒì„± ì‹œë„ (fpdf2 ì‚¬ìš©)
                        try:
                            import subprocess
                            subprocess.check_call(['pip', 'install', 'fpdf2'])
                            from fpdf import FPDF as FPDF2
                            
                            st.info("ëŒ€ì²´ ë°©ë²•ìœ¼ë¡œ PDF ìƒì„±ì„ ì‹œë„í•©ë‹ˆë‹¤...")
                            pdf = FPDF2()
                            pdf.add_page()
                            pdf.set_font('Helvetica', 'B', 16)
                            pdf.cell(40, 10, 'ë¹„ìƒì¥ì£¼ì‹ í‰ê°€ ê²°ê³¼')
                            pdf_bytes = pdf.output()
                            
                            st.success("PDF ìƒì„± ì™„ë£Œ! ì•„ë˜ ë²„íŠ¼ì„ í´ë¦­í•˜ì—¬ ë‹¤ìš´ë¡œë“œí•˜ì„¸ìš”.")
                            st.download_button(
                                label="ğŸ“„ PDF íŒŒì¼ ë‹¤ìš´ë¡œë“œ (ê¸°ë³¸ í˜•ì‹)",
                                data=pdf_bytes,
                                file_name=f"ë¹„ìƒì¥ì£¼ì‹_í‰ê°€_{st.session_state.company_name}_{st.session_state.eval_date}_ê¸°ë³¸.pdf",
                                mime="application/pdf"
                            )
                        except:
                            st.error("PDF ìƒì„±ì— ì‹¤íŒ¨í–ˆìŠµë‹ˆë‹¤. HTML í˜•ì‹ìœ¼ë¡œ ë‹¤ìš´ë¡œë“œí•´ë³´ì„¸ìš”.")
                            st.info("ë˜ëŠ” 'pip install fpdf fpdf2' ëª…ë ¹ìœ¼ë¡œ í•„ìš”í•œ ë¼ì´ë¸ŒëŸ¬ë¦¬ë¥¼ ì„¤ì¹˜í•´ë³´ì„¸ìš”.")
        
        with tab2:
            if st.button("HTML íŒŒì¼ ìƒì„±í•˜ê¸°", key="generate_html"):
                html_content = create_html_content()
                st.download_button(
                    label="ğŸ“„ HTML íŒŒì¼ ë‹¤ìš´ë¡œë“œ",
                    data=html_content,
                    file_name=f"ë¹„ìƒì¥ì£¼ì‹_í‰ê°€_{st.session_state.company_name}_{st.session_state.eval_date}.html",
                    mime="text/html"
                )
        
        with tab3:
            if st.button("CSV íŒŒì¼ ìƒì„±í•˜ê¸°", key="generate_csv"):
                csv_content = create_csv_content()
                if csv_content:
                    st.download_button(
                        label="ğŸ“„ CSV íŒŒì¼ ë‹¤ìš´ë¡œë“œ",
                        data=csv_content,
                        file_name=f"ë¹„ìƒì¥ì£¼ì‹_í‰ê°€_{st.session_state.company_name}_{st.session_state.eval_date}.csv",
                        mime="text/csv"
                    )
        st.markdown("</div>", unsafe_allow_html=True)
